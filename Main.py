import tkinter as tk
from tkinter import messagebox, font, ttk
import sqlite3
from openpyxl import Workbook, load_workbook
from datetime import datetime
import os

# Database Setup
def connect_db():
    conn = sqlite3.connect("inventory.db")
    cursor = conn.cursor()
    cursor.execute("""CREATE TABLE IF NOT EXISTS users (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT UNIQUE NOT NULL,
            password TEXT NOT NULL,
            role TEXT NOT NULL CHECK(role IN ('Admin', 'User'))
        )""")
    cursor.execute("""CREATE TABLE IF NOT EXISTS products (
            product_id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT NOT NULL,
            category TEXT NOT NULL,
            price REAL NOT NULL,
            stock_quantity INTEGER NOT NULL
        )""")
    try:
        cursor.execute("INSERT INTO users (username, password, role) VALUES ('admin', 'admin123', 'Admin')")
        cursor.execute("INSERT INTO users (username, password, role) VALUES ('user', 'user123', 'User')")
    except sqlite3.IntegrityError:
        pass
    conn.commit()
    return conn

# User Authentication
def authenticate_user(username, password):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT role FROM users WHERE username=? AND password=?", (username, password))
    user = cursor.fetchone()
    conn.close()
    return user[0] if user else None

# Product Management
def add_product(name, category, price, stock, tree):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("INSERT INTO products (name, category, price, stock_quantity) VALUES (?, ?, ?, ?)",
                   (name, category, float(price), int(stock)))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", "Product added successfully")
    log_to_excel(name, category, price, stock)
    refresh_product_table(tree)

def update_stock(product_id, new_stock, tree):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("UPDATE products SET stock_quantity=? WHERE product_id=?", (new_stock, product_id))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", "Stock updated successfully")
    refresh_product_table(tree)

def delete_product(product_id, tree):
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("DELETE FROM products WHERE product_id=?", (product_id,))
    conn.commit()
    conn.close()
    messagebox.showinfo("Success", "Product deleted successfully")
    refresh_product_table(tree)

def get_products():
    conn = connect_db()
    cursor = conn.cursor()
    cursor.execute("SELECT * FROM products")
    products = cursor.fetchall()
    conn.close()
    return products

def refresh_product_table(tree):
    for item in tree.get_children():
        tree.delete(item)
    products = get_products()
    for product in products:
        tree.insert("", "end", values=product)

# Excel Logging
def log_to_excel(name, category, price, stock):
    filename = "product_log.xlsx"
    file_exists = os.path.isfile(filename)
    if not file_exists:
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "Product Log"
        sheet.append(["Date", "Product Name", "Category", "Price", "Stock Quantity"])
    else:
        workbook = load_workbook(filename)
        sheet = workbook.active
    current_date = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    sheet.append([current_date, name, category, float(price), int(stock)])
    workbook.save(filename)

# Billing
def generate_bill(items):
    total = sum(item['price'] * item['quantity'] for item in items)
    bill_window = tk.Toplevel()
    bill_window.title("Generated Bill")

    columns = ("Item", "Quantity", "Price", "Total")
    tree = ttk.Treeview(bill_window, columns=columns, show='headings')
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center")
    tree.pack(expand=True, fill="both", padx=10, pady=10)

    for item in items:
        total_price = item['price'] * item['quantity']
        tree.insert("", "end", values=(item['name'], item['quantity'], f"${item['price']:.2f}", f"${total_price:.2f}"))

    tk.Label(bill_window, text=f"Total Amount: ${total:.2f}", font=("Arial", 14), fg="blue").pack(pady=10)

# Admin Panel
def create_admin_panel(root):
    root.title("Admin Panel")
    root.geometry("800x600")
    notebook = ttk.Notebook(root)

    # Add Products Tab
    add_tab = ttk.Frame(notebook)
    notebook.add(add_tab, text="Add Products")
    setup_add_product_tab(add_tab)

    # View Products Tab
    view_tab = ttk.Frame(notebook)
    notebook.add(view_tab, text="View Products")
    tree = setup_product_view_tab(view_tab)
    
    notebook.pack(expand=True, fill="both", padx=10, pady=10)
    ttk.Button(root, text="Logout", command=lambda: logout(root), width=10).pack(pady=10)

def setup_add_product_tab(tab):
    tk.Label(tab, text="Add New Product", font=("Arial", 14)).pack(pady=10)
    entries = {}
    for label in ["Product Name", "Category", "Price", "Stock Quantity"]:
        tk.Label(tab, text=label).pack(pady=5)
        entry = tk.Entry(tab, width=30)
        entry.pack()
        entries[label] = entry

    add_button = ttk.Button(tab, text="Add Product", command=lambda: add_product(
        entries["Product Name"].get(),
        entries["Category"].get(),
        entries["Price"].get(),
        entries["Stock Quantity"].get(),
        tab.master.master.children['!frame']))
    add_button.pack(pady=10)

def setup_product_view_tab(tab):
    columns = ("Product ID", "Name", "Category", "Price", "Stock Quantity")
    tree = ttk.Treeview(tab, columns=columns, show='headings')
    for col in columns:
        tree.heading(col, text=col)
        tree.column(col, anchor="center")
    tree.pack(expand=True, fill="both", padx=10, pady=10)
    refresh_product_table(tree)
    return tree

# User Panel
def create_user_panel(root):
    root.title("User Panel")
    root.geometry("800x600")
    ttk.Button(root, text="Logout", command=lambda: logout(root), width=10).pack(pady=10)

# Login Page
def create_login_page(root):
    root.title("Login")
    root.geometry("400x300")
    heading_font = font.Font(family="Bebas Neue", size=24, weight="bold")
    tk.Label(root, text="User Login", font=heading_font).pack(pady=20)
    
    username_entry = tk.Entry(root)
    password_entry = tk.Entry(root, show="*")

    username_entry.pack(pady=10)
    password_entry.pack(pady=10)
    
    ttk.Button(root, text="Login", command=lambda: login_action(root, username_entry.get(), password_entry.get())).pack(pady=10)

def login_action(root, username, password):
    role = authenticate_user(username, password)
    if role:
        root.destroy()
        new_root = tk.Tk()
        if role == 'Admin':
            create_admin_panel(new_root)
        else:
            create_user_panel(new_root)
        new_root.mainloop()
    else:
        messagebox.showerror("Login Failed", "Invalid username or password.")

def logout(root):
    root.destroy()
    new_root = tk.Tk()
    create_login_page(new_root)
    new_root.mainloop()

# Main Function
if __name__ == "__main__":
    root = tk.Tk()
    create_login_page(root)
    root.mainloop()
